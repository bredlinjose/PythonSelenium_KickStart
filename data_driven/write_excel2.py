import os

import openpyxl

# file --> workbook --> sheets --> rows --> cells
file = os.path.dirname(os.path.abspath('.')) + '\\files' + "\\test_data.xlsx"
workbook = openpyxl.open(file)
sheet = workbook["Sheet2"]

sheet.cell(1, 1).value = 150
sheet.cell(1, 2).value = 'Smith'

sheet.cell(2, 1).value = 151
sheet.cell(2, 2).value = 'John'

sheet.cell(3, 1).value = 152
sheet.cell(3, 2).value = 'Adam'

workbook.save(file)

