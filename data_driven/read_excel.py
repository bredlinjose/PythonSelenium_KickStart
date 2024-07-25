import os

import openpyxl

# file --> workbook --> sheets --> rows --> cells
file = os.path.dirname(os.path.abspath('.')) + '\\files' + "\\test_data.xlsx"
workbook = openpyxl.open(file)
sheet = workbook["books"]
# sheet = workbook.active  # get active sheet from Excel  #used only for single sheet

rows = sheet.max_row
print("row count:", rows)
columns = sheet.max_column
print("column count:", columns)

for r in range(1, rows+1):
    for c in range(1, columns+1):
        print(sheet.cell(r, c).value, end=' ')
    print()



