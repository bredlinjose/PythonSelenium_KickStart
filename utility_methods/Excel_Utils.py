import openpyxl
from openpyxl.styles import PatternFill


def get_row_count(file, sheetName):
    workbook = openpyxl.open(file)
    sheet = workbook[sheetName]
    return sheet.max_row


def get_column_count(file, sheetName):
    workbook = openpyxl.open(file)
    sheet = workbook[sheetName]
    return sheet.max_column


def read_data(file, sheetName, rowNum, columnNum):
    workbook = openpyxl.open(file)
    sheet = workbook[sheetName]
    return sheet.cell(rowNum, columnNum).value


def write_data(file, sheetName, rowNum, columnNum, data):
    workbook = openpyxl.open(file)
    sheet = workbook[sheetName]
    sheet.cell(rowNum, columnNum).value = data
    workbook.save(file)


def fill_green_colour(file, sheetName, rowNum, columnNum):
    workbook = openpyxl.open(file)
    sheet = workbook[sheetName]
    greenFill = PatternFill(start_color='60b212', end_color='60b212', fill_type='solid')
    sheet.cell(rowNum, columnNum).fill = greenFill
    workbook.save(file)

def fill_red_colour(file, sheetName, rowNum, columnNum):
    workbook = openpyxl.open(file)
    sheet = workbook[sheetName]
    redFill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
    sheet.cell(rowNum, columnNum).fill = redFill
    workbook.save(file)
